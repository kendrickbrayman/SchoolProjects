import glob
import openpyxl as opx
import re

def getFilenames(dirName):
    """
    :param dirName: directory name to check for filenames in
    :return: list of all file names that match the RegEx
    """
    #prompts the user to select a directory and selects all .txt files for parsing
    names = glob.glob(dirName + "/*.txt")
    return names


def text2table(text):
    """
    :param text: the table text to be converted into a table format
    :return: the finished table
    """
    out = [re.split('\t',entry) for entry in re.split('\n+',text)]
    return out

def parseLine(line,dic):
    """

    :param line: single line of text to be parsed
    :param dic: dictionary of RexEx strings to be searched
    :return: key and match if there is a match and none,none if there is no match in the string
    """
    if len(dic) == 0:
        return None,None
    for key,rx in dic.items():
        match = re.search(rx,line)
        if match:
            del dic[key] #keeps the program efficient by not iterating over data already parsed
            return key,match.group(1)
        else:
            return None,None

def parseFile(filename):
    """
    :param filename: name of the file to be parsed
    :return: a dictionary containing the organized data in tidy format
    """

    # Dictionary of all RegExs
    rxDict = {
        # general data
        'Patient': re.compile(r'Patient\t([^\n]*)\n'),
        'Birth Date': re.compile(r'Birth Date\t([^\n]*)\n'),
        'Patient ID': re.compile(r'PatientID\t([^\n]*)\n'),
        'Custom Patient ID': re.compile(r'Custom Patient ID\t([^\n]*)\n'),
        'Study Date': re.compile(r'Study Date\t([^\n]*)\n'),
        'AccessionNumber': re.compile(r'AccessionNumber\t([^\n]*)\n'),
        'cvi42 Version': re.compile(r'cvi42 Version\t([^\n]*)\n'),
        'Measurement Exclude Zero Values': re.compile(r'Measurement Exclude Zero Values\t([^\n]*)\n'),
        'Offset Endo': re.compile(r'Offset Endo\t([^\n]*)%\n'),
        'Offset Epi': re.compile(r'Offset Epi\t([^\n]*)%\n'),
        # Native only data
        #Different sections for CA and Native so that they could remain on the single line search portion of the function
        'Ncvi42 Series ID': re.compile(r'cvi42 Series ID\t([^\n]*)\n'),
        'NModality': re.compile(r'Modality\t([^\n]*)', flags=re.DOTALL),
        'NSeries Number(s)': re.compile(r'Series Number\(s\)\t([^\n]*)\n', flags=re.DOTALL),
        'NSeries Description': re.compile(r'Series Description\t([^\n]*)\n', flags=re.DOTALL),
        'NSeries Time': re.compile(r'Series Time\t([^\n]*)\n', flags=re.DOTALL),
        'NSequence Name': re.compile(r'Sequence Name\t([^\n]*)\n', flags=re.DOTALL),
        'NProtocol Name': re.compile(r'Protocol Name\t([^\n]*)\n', flags=re.DOTALL),
        'NManufacturer': re.compile(r'Manufacturer\t([^\n]*)\n', flags=re.DOTALL),
        'NGlobal Myo T1 Across Slices': re.compile(r'Global Myo T1 Across Slices\t([^\n]*)\n', flags=re.DOTALL),
        # CA only data
        'Ccvi42 Series ID': re.compile(r'cvi42 Series ID\t([^\n]*)\n'),
        'CModality': re.compile(r'Modality\t([^\n]*)', flags=re.DOTALL),
        'CSeries Number(s)': re.compile(r'Series Number\(s\)\t([^\n]*)\n', flags=re.DOTALL),
        'CSeries Description': re.compile(r'Series Description\t([^\n]*)\n', flags=re.DOTALL),
        'CSeries Time': re.compile(r'Series Time\t([^\n]*)\n', flags=re.DOTALL),
        'CSequence Name': re.compile(r'Sequence Name\t([^\n]*)\n', flags=re.DOTALL),
        'CProtocol Name': re.compile(r'Protocol Name\t([^\n]*)\n', flags=re.DOTALL),
        'CManufacturer': re.compile(r'Manufacturer\t([^\n]*)\n', flags=re.DOTALL),
        'CGlobal Myo T1 Across Slices': re.compile(r'Global Myo T1 Across Slices\t([^\n]*)\n', flags=re.DOTALL)
    }
    # dictionary for multiline RegEx searches
    rxMultiDict = {
        'Mean T1(ms)': '(?:Native|CA) T1 \(AHA.*?(1\t[\d]*.*?16\t[^\n\s]*)',
        'Slice Data': '(?:Native|CA) T1 Slice [123].*?(1\t.*?)\n\n',
    }


    #reads single line data into parsed dictionary
    with open(filename,'r') as f_obj:
        line = re.sub('\0','',f_obj.readline())
        parsed = {}
        while line:
            key,match = parseLine(line,rxDict)
            parsed[key] = match
            line = re.sub('\0','',f_obj.readline()) #strips the null characters
        f_obj.close()

    #reads data from tables into parsed dictionary
    with open(filename, 'r') as f_obj:
        rawTxt = re.sub('\0','',f_obj.read())
        for key,rx in rxMultiDict.items():
            parsed[key] = re.findall(rx,rawTxt,flags=re.DOTALL)

    #takes text tables and converts them to lists

    for i in range(6):
        parsed['Slice Data'][i] = [text2table(parsed['Slice Data'][i])]

    #compiles the master dictionary to output
    masterData = {
        'Patient Info': {
            'Patient': parsed['Patient'],
            'Birth Date': parsed['Birth Date'],
            'Patient ID': parsed['Patient ID'],
            'Custom Patient ID': parsed['Custom Patient ID'],
            'Study Date': re.split('\t',parsed['Study Date'])[0],
            'Study Time': re.split('\t', parsed['Study Date'])[1],
            'Accession Number': parsed['AccessionNumber'],
            'cvi42 Version': parsed['cvi42 Version'],
            'Measurement Exclude Zero Values': parsed['Measurement Exclude Zero Values'],
            'Offset Endo': float(parsed['Offset Endo'])/100,
            'Offset Epi': float(parsed['Offset Epi'])/100
        },
        'T1 Info': {
            'Patient ID': [parsed['Patient']] * 2,
            'Native / CA': ['Native','CA'],
            'cvi42 Series ID': [parsed['Ncvi42 Series ID'],parsed['Ccvi42 Series ID']],
            'Modality': [parsed['NModality'],parsed['CModality']],
            'Series Number(s)': [parsed['NSeries Number(s)'],parsed['CSeries Number(s)']],
            'Series Description': [parsed['NSeries Description'],parsed['CSeries Description']],
            'Series Time': [parsed['NSeries Time'],parsed['CSeries Time']],
            'Sequence Name': [parsed['NSequence Name'],parsed['CSequence Name']],
            'Protocol Name': [parsed['NProtocol Name'],parsed['CProtocol Name']],
            'Manufacturer': [parsed['NManufacturer'],parsed['CManufacturer']],
            'Global Myo T1 Across Slices': [parsed['NGlobal Myo T1 Across Slices'],parsed['CGlobal Myo T1 Across Slices']]
        },
        'Regional T1': {
            'Patient ID': [parsed['Patient']] * 16 * 2,
            'Native / CA': ['Native'] * 16 + ['CA'] * 16,
            'Segment': [text2table(parsed['Mean T1(ms)'][j])[i][0] for j in range(2) for i in range(16)], #kept as string intentionally to indicate factor
            'Mean T1(ms)': [float(text2table(parsed['Mean T1(ms)'][j])[i][1]) for j in range(2) for i in range(16)]
        },
        'Regional T1 Slice': {
            'Patient ID': [parsed['Patient']] * 48,
            'Native / CA': ['Native'] * 8 * 3 + ['CA'] * 8 * 3,
            'Slice': (['1'] * 8 + ['2'] * 8 + ['3'] * 8) * 2, #kept as string intentionally to indicate factor
            'Segment / Region': [parsed['Slice Data'][i][0][j][0] for i  in range(6) for j in range(8)], #kept as string intentionally to indicate factor
            'Mean(ms)': [float(parsed['Slice Data'][i][0][j][1]) for i  in range(6) for j in range(8)],
            'Median(ms)': [float(parsed['Slice Data'][i][0][j][2]) for i  in range(6) for j in range(8)],
            'Min(ms)': [float(parsed['Slice Data'][i][0][j][3]) for i  in range(6) for j in range(8)],
            'Max(ms)':[float(parsed['Slice Data'][i][0][j][4]) for i  in range(6) for j in range(8)],
            'SD(ms)': [float(parsed['Slice Data'][i][0][j][5]) for i  in range(6) for j in range(8)],
            'Area(mm ^ 2)': [float(parsed['Slice Data'][i][0][j][6]) for i  in range(6) for j in range(8)]
        }
    }
    return masterData



def createWorkbook(wkbName):
    """
    :param wkbName: the string name for the workbook to be created. Creates a workbook in the current directory if no file path is passed as part of the name.
    NOTE: this param requires a .xlsx ending
    :return: the workbook name
    """
    book = opx.Workbook()
    #initializes the sheets to add the data to

    book['Sheet'].title = 'Patient Info'
    #book.remove_sheet('Sheet1')

    book['Patient Info'].append(("Patient","Birth Date","Patient ID","Custom Patient ID","Study Date","Study Time","Accession Number","cvi42 Version","Measurement Exclude Zero Values","Offset Endo","Offset Epi"))
    book.create_sheet('T1 Info')
    book['T1 Info'].append(("Patient ID","Native/CA","cvi42 Series ID","Modality","Series Number(s)","Series Desciption","Series Time","Sequence Name","Protocol Name","Manufacturer","Global Myo T1 Across Slices"))
    book.create_sheet('Regional T1')
    book['Regional T1'].append(("Patient ID","Native/CA","Segment", "Mean T1 (ms)"))
    book.create_sheet('Regional T1 Slice')
    book['Regional T1 Slice'].append(("Patient ID","Native/CA","Slice","Segment/Region","Mean (ms)","Median (ms)","Min (ms)","Max (ms)","SD (ms)","Area (mm^2)"))
    #initializes the headers for the different worksheets
    book.save(wkbName)
    return wkbName

def updateWorkbook(wkbkName,data):
    """
    :param wkbkName: File path + file name for the workbook being updated
    :param data:
    :return: 0 for a sucessful workbook update/save
    """
    book = opx.load_workbook(wkbkName)
    for sheet,data in data.items():
        active = book[sheet]
        # checks for the data having multiple entries per patient for example tabular data over metadata
        if type(tuple(data.values())[0]) is str:
            active.append(tuple(data.values()))
        else:
            for i in range(len(tuple(data.values())[0])):
                active.append(tuple([tuple(data.values())[j][i] for j in range(len(data.values()))]))

    book.save(wkbkName)
    return 0


def main():
    print("Enter directory to check for files to import:")
    searchDir = input()
    fileNames = getFilenames(searchDir)

    patients = []
    for name in fileNames:
        patientData = parseFile(name)
        patients.append(patientData)

    print("Enter output file name (data will append if filename already exists):")  # file should be saved in same dir as searchDir
    workbookName = searchDir + "\\" + input()
    #tests if the script is appending to an existing workbook and creates a new one if there is no existing one
    try:
        opx.load_workbook(workbookName)
    except:
        createWorkbook(workbookName)

    for data in patients:
        updateWorkbook(workbookName, data)


if __name__ == '__main__':
    main()
